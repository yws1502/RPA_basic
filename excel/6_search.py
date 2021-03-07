from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active

# 컬럼명 확인하기
# for col in tuple(ws.iter_cols()):
#     print(col[0].value)

# 번호 영어 수학
for row in tuple(ws.iter_rows(min_row=2)):
    if row[1].value > 80:
        print(row[0].value, '번 영잘알')

for col in tuple(ws.iter_cols(max_row=1)):
    for cell in col:
        if cell.value == '영어':
            cell.value = '컴퓨터'



wb.save('sample_modified.xlsx')